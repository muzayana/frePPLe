#
# Copyright (C) 2007-2013 by frePPLe bvba
#
# All information contained herein is, and remains the property of frePPLe.
# You are allowed to use and modify the source code, as long as the software is used
# within your company.
# You are not allowed to distribute the software, either in the form of source code
# or in the form of compiled binaries.
#

r'''
This module implements a generic view to presents lists and tables.

It provides the following functionality:
 - Pagination of the results.
 - Ability to filter on fields, using different operators.
 - Ability to sort on a field.
 - Export the results as a CSV file, ready for use in a spreadsheet.
 - Import CSV formatted data files.
 - Show time buckets to show data by time buckets.
   The time buckets and time boundaries can easily be updated.
'''

import codecs
import collections
import csv
from datetime import datetime, timedelta
from decimal import Decimal
import functools
import math
import operator
import json
from io import StringIO, BytesIO
from openpyxl import load_workbook, Workbook

from django.apps import apps
from django.contrib.auth.models import Group
from django.contrib.auth import get_permission_codename
from django.conf import settings
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.admin.utils import unquote, quote
from django.core.exceptions import ValidationError
from django.core.management.color import no_style
from django.db import connections, transaction, models
from django.db.models.fields import Field, CharField, IntegerField, AutoField
from django.db.models.fields.related import RelatedField
from django.forms.models import modelform_factory
from django.http import Http404, HttpResponse, StreamingHttpResponse
from django.http import HttpResponseForbidden, HttpResponseNotAllowed
from django.shortcuts import render
from django.utils import translation, six
from django.utils.decorators import method_decorator
from django.utils.encoding import smart_str, iri_to_uri, force_text
from django.utils.html import escape
from django.utils.translation import ugettext as _
from django.utils.formats import get_format
from django.utils.text import capfirst, get_text_list
from django.utils.translation import string_concat
from django.template.defaultfilters import title
from django.contrib.admin.models import LogEntry, CHANGE, ADDITION, DELETION
from django.contrib.contenttypes.models import ContentType
from django.views.generic.base import View


from freppledb.common.models import User, Comment, Parameter, BucketDetail, Bucket, HierarchyModel


import logging
logger = logging.getLogger(__name__)


# A list of models with some special, administrative purpose.
# They should be excluded from bulk import, export and erasing actions.
EXCLUDE_FROM_BULK_OPERATIONS = (Group, User, Comment)


class GridField(object):
  '''
  Base field for columns in grid views.
  '''

  def __init__(self, name, **kwargs):
    self.name = name
    for key, value in kwargs.items():
      setattr(self, key, value)
    if 'key' in kwargs:
      self.editable = False
    if 'title' not in kwargs and not self.title:
      self.title = self.name and _(self.name) or ''
    if not self.name:
      self.sortable = False
      self.search = False
    if 'field_name' not in kwargs:
      self.field_name = self.name

  def __str__(self):
    o = [ "name:'%s',index:'%s',editable:%s,label:'%s',align:'%s',title:false" %
          (self.name or '', self.name or '', self.editable and "true" or "false",
           force_text(self.title).title().replace("'", "\\'"), self.align
           ), ]
    if self.key:
      o.append( ",key:true" )
    if not self.sortable:
      o.append(",sortable:false")
    if not self.search:
      o.append(",search:false")
    if self.formatter:
      o.append(",formatter:'%s'" % self.formatter)
    if self.unformat:
      o.append(",unformat:'%s'" % self.unformat)
    if self.searchrules:
      o.append(",searchrules:{%s}" % self.searchrules)
    if self.hidden:
      o.append(",hidden:true")
    if self.extra:
      o.append(",%s" % force_text(self.extra))
    return ''.join(o)

  name = None
  field_name = None
  formatter = None
  width = 100
  editable = True
  sortable = True
  search = True
  key = False
  unformat = None
  title = None
  extra = None
  align = 'center'
  searchrules = None
  hidden = False


class GridFieldDateTime(GridField):
  formatter = 'date'
  extra = "formatoptions:{srcformat:'Y-m-d H:i:s',newformat:'Y-m-d H:i:s'}"
  width = 140


class GridFieldTime(GridField):
  formatter = 'time'
  extra = "formatoptions:{srcformat:'H:i:s',newformat:'H:i:s'}"
  width = 80


class GridFieldDate(GridField):
  formatter = 'date'
  extra = "formatoptions:{srcformat:'Y-m-d',newformat:'Y-m-d'}"
  width = 140


class GridFieldInteger(GridField):
  formatter = 'integer'
  width = 70
  searchrules = 'integer:true'


class GridFieldNumber(GridField):
  formatter = 'number'
  width = 70
  searchrules = 'number:true'


class GridFieldBool(GridField):
  extra = "formatoptions:{disabled:false}, edittype:'checkbox', editoptions:{value:'True:False'}"
  width = 60


class GridFieldLastModified(GridField):
  formatter = 'date'
  extra = "formatoptions:{srcformat:'Y-m-d H:i:s',newformat:'Y-m-d H:i:s'}"
  title = _('last modified')
  editable = False
  width = 140


class GridFieldText(GridField):
  width = 200
  align = 'left'


class GridFieldChoice(GridField):
  width = 100
  align = 'center'

  def __init__(self, name, **kwargs):
    super(GridFieldChoice, self).__init__(name, **kwargs)
    e = ["formatter:'select', edittype:'select', editoptions:{value:'"]
    first = True
    for i in kwargs["choices"]:
      if first:
        first = False
        e.append("%s:" % i[0])
      else:
        e.append(";%s:" % i[0])
      e.append(i[1])
    e.append("'}")
    self.extra = string_concat(*e)


class GridFieldCurrency(GridField):
  formatter = 'currency'
  extra = "formatoptions:{prefix:'%s', suffix:'%s'}" % settings.CURRENCY
  width = 80


class GridFieldDuration(GridField):
  formatter = 'duration'
  width = 80


def getBOM(encoding):
  try:
    # Get the official name of the encoding (since encodings can have many alias names)
    name = codecs.lookup(encoding).name
  except:
    return ''  # Unknown encoding, without BOM header
  if name == 'utf-32-be':
    return codecs.BOM_UTF32_BE
  elif name == 'utf-32-le':
    return codecs.BOM_UTF32_LE
  elif name == 'utf-16-be':
    return codecs.BOM_UTF16_BE
  elif name == 'utf-16-le':
    return codecs.BOM_UTF16_LE
  elif name == 'utf-8':
    return codecs.BOM_UTF8
  else:
    return ''


class EncodedCSVReader:
  """
  A CSV reader which will iterate over lines in the CSV data buffer.
  The reader will scan the BOM header in the data to detect the right encoding.
  """
  def __init__(self, datafile, **kwds):
    # Read the file into memory
    # TODO Huge file uploads can overwhelm your system!
    data = datafile.read()
    # Detect the encoding of the data by scanning the BOM.
    # Skip the BOM header if it is found.
    if data.startswith(codecs.BOM_UTF32_BE):
      self.reader = StringIO(data.decode('utf_32_be'))
      self.reader.read(1)
    elif data.startswith(codecs.BOM_UTF32_LE):
      self.reader = StringIO(data.decode('utf_32_le'))
      self.reader.read(1)
    elif data.startswith(codecs.BOM_UTF16_BE):
      self.reader = StringIO(data.decode('utf_16_be'))
      self.reader.read(1)
    elif data.startswith(codecs.BOM_UTF16_LE):
      self.reader = StringIO(data.decode('utf_16_le'))
      self.reader.read(1)
    elif data.startswith(codecs.BOM_UTF8):
      self.reader = StringIO(data.decode('utf_8'))
      self.reader.read(1)
    else:
      # No BOM header found. We assume the data is encoded in the default CSV character set.
      self.reader = StringIO(data.decode(settings.CSV_CHARSET))
    self.csvreader = csv.reader(self.reader, **kwds)

  def __next__(self):
    return next(self.csvreader)

  def __iter__(self):
    return self


class GridReport(View):
  '''
  The base class for all jqgrid views.
  The parameter values defined here are used as defaults for all reports, but
  can be overwritten.
  '''
  # Points to template to be used
  template = 'admin/base_site_grid.html'

  # The title of the report. Used for the window title
  title = ''

  # The resultset that returns a list of entities that are to be
  # included in the report.
  # This query is used to return the number of records.
  # It is also used to generate the actual results, in case no method
  # "query" is provided on the class.
  basequeryset = None

  # Specifies which column is used for an initial ordering
  default_sort = (0, 'asc')

  # A model class from which we can inherit information.
  model = None

  # Which admin site is used for the model: 'data' or 'admin'
  adminsite = 'data'

  # Allow editing in this report or not
  editable = True

  # Allow filtering of the results or not
  filterable = True

  # Include time bucket support in the report
  hasTimeBuckets = False

  # Specify a minimum level for the time buckets available in the report.
  # Higher values (ie more granular) buckets can then not be selected.
  maxBucketLevel = None

  # Show a select box in front to allow selection of records
  multiselect = True

  # Control the height of the grid. By default the full browser window is used.
  height = None

  # Number of columns frozen in the report
  frozenColumns = 0

  # A list with required user permissions to view the report
  permissions = ()

  # Defines the difference between height of the grid and its boundaries
  heightmargin = 75

  # Define a list of actions
  actions = None

  @classmethod
  def getKey(cls):
    return "%s.%s" % (cls.__module__, cls.__name__)


  @classmethod
  def getAppLabel(cls):
    '''
    Return the name of the Django application which defines this report.
    '''
    if hasattr(cls, 'app_label'):
      return cls.app_label
    s = cls.__module__.split('.')
    for i in range(len(s), 0, -1):
      x = '.'.join(s[0:i])
      if x in settings.INSTALLED_APPS:
        cls.app_label = s[i - 1]
        return cls.app_label
    raise Exception("Can't identify app of reportclass %s" % cls)


  # Extra variables added to the report template
  @classmethod
  def extra_context(reportclass, request, *args, **kwargs):
    return {}


  @classmethod
  def getBuckets(reportclass, request, *args, **kwargs):
    '''
    This function gets passed a name of a bucketization.
    It returns a tuple with:
      - the start date of the report horizon
      - the end date of the reporting horizon
      - a list of buckets.
    '''
    # Pick up the user preferences
    pref = request.user

    # Select the bucket size (unless it is passed as argument)
    try:
      bucket = Bucket.objects.using(request.database).get(name=pref.horizonbuckets)
    except:
      try:
        bucket = Bucket.objects.using(request.database).order_by('-level')[0].name
      except:
        bucket = None

    if pref.horizontype:
      # First type: Start and end dates relative to current
      try:
        start = datetime.strptime(
          Parameter.objects.using(request.database).get(name="currentdate").value,
          "%Y-%m-%d %H:%M:%S"
          )
      except:
        start = datetime.now()
      start = start.replace(hour=0, minute=0, second=0, microsecond=0)
      if pref.horizonunit == 'day':
        end = start + timedelta(days=pref.horizonlength or 60)
        end = end.replace(hour=0, minute=0, second=0)
      elif pref.horizonunit == 'week':
        end = start.replace(hour=0, minute=0, second=0) + timedelta(weeks=pref.horizonlength or 8, days=7 - start.weekday())
      else:
        y = start.year
        m = start.month + (pref.horizonlength or 2) + (start.day > 1 and 1 or 0)
        while m > 12:
          y += 1
          m -= 12
        end = datetime(y, m, 1)
    else:
      # Second type: Absolute start and end dates given
      start = pref.horizonstart
      if not start:
        try:
          start = datetime.strptime(
            Parameter.objects.using(request.database).get(name="currentdate").value,
            "%Y-%m-%d %H:%M:%S"
            )
        except:
          start = datetime.now()
          start = start.replace(microsecond=0)
      end = pref.horizonend
      if end:
        if end < start:
          # Swap start and end to assure the start is before the end
          tmp = start
          start = end
          end = tmp
      else:
        if pref.horizonunit == 'day':
          end = start + timedelta(days=pref.horizonlength or 60)
        elif pref.horizonunit == 'week':
          end = start + timedelta(weeks=pref.horizonlength or 8)
        else:
          end = start + timedelta(weeks=pref.horizonlength or 8)

    # Filter based on the start and end date
    request.report_startdate = start
    request.report_enddate = end
    request.report_bucket = str(bucket)
    if bucket:
      res = BucketDetail.objects.using(request.database).filter(bucket=bucket)
      if start:
        res = res.filter(enddate__gt=start)
      if end:
        res = res.filter(startdate__lt=end)
      request.report_bucketlist = res.values('name', 'startdate', 'enddate')
    else:
      request.report_bucketlist = []


  @method_decorator(staff_member_required)
  @method_decorator(csrf_protect)
  def dispatch(self, request, *args, **kwargs):
    # Verify the user is authorized to view the report
    for perm in self.permissions:
      if not request.user.has_perm("%s.%s" % (self.getAppLabel(), perm[0])):
        return HttpResponseForbidden('<h1>%s</h1>' % _('Permission denied'))

    # Unescape special characters in the arguments.
    # All arguments are encoded with escaping function used on the django admin.
    args_unquoted = [ unquote(i) for i in args ]

    # Dispatch to the correct method
    if request.method == 'GET':
      return self.get(request, *args_unquoted, **kwargs)
    elif request.method == 'POST':
      return self.post(request, *args_unquoted, **kwargs)
    else:
      return HttpResponseNotAllowed(['get', 'post'])


  @classmethod
  def _render_colmodel(cls, is_popup=False, prefs=None, mode="graph"):
    if not prefs:
      frozencolumns = cls.frozenColumns
      rows = [ (i, cls.rows[i].hidden, cls.rows[i].width) for i in range(len(cls.rows)) ]
    else:
      frozencolumns = prefs.get('frozen', cls.frozenColumns)
      rows = prefs.get('rows')
      if not rows:
        rows = [ (i, cls.rows[i].hidden, cls.rows[i].width) for i in range(len(cls.rows)) ]
    result = []
    if is_popup:
      result.append("{name:'select',label:gettext('Select'),width:75,align:'center',sortable:false,search:false}")
    count = -1
    for (index, hidden, width) in rows:
      count += 1
      try:
        result.append("{%s,width:%s,counter:%d%s%s%s,searchoptions:{searchhidden: true}}" % (
           cls.rows[index], width, index,
           count < frozencolumns and ',frozen:true' or '',
           is_popup and ',popup:true' or '',
           hidden and not cls.rows[index].hidden and ',hidden:true' or ''
           ))
      except IndexError:
        logger.warning('Invalid preference value for %s: %s' % (cls.getKey(), prefs))
    return ',\n'.join(result)


  @classmethod
  def _generate_spreadsheet_data(reportclass, request, *args, **kwargs):
    # Create a workbook
    wb = Workbook(optimized_write=True)
    title = force_text(reportclass.model and reportclass.model._meta.verbose_name or reportclass.title)
    ws = wb.create_sheet(title=title)

    # Choose fields to export and write the title row
    prefs = request.user.getPreference(reportclass.getKey())
    if prefs:
      prefs = prefs.get('rows', None)
    if prefs:
      # Customized settings
      fields = [
        reportclass.rows[f[0]]
        for f in prefs
        if not f[1] and f[0] < len(reportclass.rows) and not reportclass.rows[f[0]].hidden
        ]
    else:
      # Default settings
      fields = [ i for i in reportclass.rows if i.field_name and not i.hidden ]
    field_names = [ f.field_name for f in fields]

    # Write a header row
    ws.append([ force_text(f.title).title() for f in fields ])

    # Loop over all records
    if isinstance(reportclass.basequeryset, collections.Callable):
      query = reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database))
    else:
      query = reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset).using(request.database))
    for row in hasattr(reportclass, 'query') and reportclass.query(request, query) or query.values(*field_names):
      if hasattr(row, "__getitem__"):
        ws.append([ _getCellValue(row[f]) for f in field_names ])
      else:
        ws.append([ _getCellValue(getattr(row, f)) for f in field_names ])

    # Write the spreadsheet from memory to a string and then to a HTTP response
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      content=output.getvalue()
      )
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % title
    response['Cache-Control'] = "no-cache, no-store"
    return response


  @classmethod
  def _generate_csv_data(reportclass, request, *args, **kwargs):
    sf = StringIO()
    decimal_separator = get_format('DECIMAL_SEPARATOR', request.LANGUAGE_CODE, True)
    if decimal_separator == ",":
      writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=';')
    else:
      writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')
    if translation.get_language() != request.LANGUAGE_CODE:
      translation.activate(request.LANGUAGE_CODE)

    # Write a Unicode Byte Order Mark header, aka BOM (Excel needs it to open UTF-8 file properly)
    encoding = settings.CSV_CHARSET
    yield getBOM(encoding)

    # Choose fields to export
    prefs = request.user.getPreference(reportclass.getKey())
    if prefs:
      prefs = prefs.get('rows', None)
    if prefs:
      # Customized settings
      prefs = prefs['rows']
      writer.writerow([
        force_text(reportclass.rows[f[0]].title, encoding=encoding, errors="ignore").title()
        for f in prefs
        if not f[1] and f[0] < len(reportclass.rows) and not reportclass.rows[f[0]].hidden
        ])
      fields = [
        reportclass.rows[f[0]].field_name
        for f in prefs
        if not f[1] and f[0] < len(reportclass.rows) and not reportclass.rows[f[0]].hidden
        ]
    else:
      # Default settings
      writer.writerow([
        force_text(f.title, encoding=encoding, errors="ignore").title()
        for f in reportclass.rows
        if f.title and not f.hidden
        ])
      fields = [
        i.field_name
        for i in reportclass.rows
        if i.field_name and not i.hidden
        ]

    # Write a header row
    yield sf.getvalue()

    # Write the report content
    if isinstance(reportclass.basequeryset, collections.Callable):
      query = reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database), prefs)
    else:
      query = reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset).using(request.database), prefs)
    for row in hasattr(reportclass, 'query') and reportclass.query(request, query) or query.values(*fields):
      # Clear the return string buffer
      sf.seek(0)
      sf.truncate(0)
      # Build the return value, encoding all output
      if hasattr(row, "__getitem__"):
        writer.writerow([
          force_text(_localize(row[f], decimal_separator), encoding=encoding, errors='ignore') if row[f] is not None else ''
          for f in fields
          ])
      else:
        writer.writerow([
          force_text(_localize(getattr(row, f), decimal_separator), encoding=encoding, errors='ignore') if getattr(row, f) is not None else ''
          for f in fields
          ])
      # Return string
      yield sf.getvalue()


  @classmethod
  def _apply_sort(reportclass, request, query, prefs=None):
    '''
    Applies a sort to the query.
    '''
    asc = True
    sort = None
    if 'sidx' in request.GET:
      sort = request.GET['sidx']
      if 'sord' in request.GET and request.GET['sord'] == 'desc':
        asc = False
    if not sort:
      if prefs and 'sidx' in prefs:
        sort = prefs['sidx']
        if 'sord' in prefs and prefs['sord'] == 'desc':
          asc = False
      if not sort and reportclass.default_sort:
        sort = reportclass.rows[reportclass.default_sort[0]].name
        if reportclass.default_sort[1] == 'desc':
          asc = False
      else:
        # No sorting
        return query
    if sort and reportclass.model:
      # Validate the field does exist.
      for name in reportclass.model._meta.get_all_field_names():
        if name == sort:
          return query.order_by(asc and sort or ('-%s' % sort))
    # Sorting by a non-existent field name: ignore the filter
    return query


  @classmethod
  def get_sort(reportclass, request):
    try:
      if 'sidx' in request.GET:
        sort = 1
        ok = False
        for r in reportclass.rows:
          if r.name == request.GET['sidx']:
            ok = True
            break
          sort += 1
        if not ok:
          sort = reportclass.default_sort[0]
      else:
        sort = reportclass.default_sort[0]
    except:
      sort = reportclass.default_sort[0]
    if ('sord' in request.GET and request.GET['sord'] == 'desc') or reportclass.default_sort[1] == 'desc':
      return "%s asc" % sort
    else:
      return "%s desc" % sort


  @classmethod
  def _generate_json_data(reportclass, request, *args, **kwargs):
    page = 'page' in request.GET and int(request.GET['page']) or 1
    if isinstance(reportclass.basequeryset, collections.Callable):
      query = reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database)
    else:
      query = reportclass.filter_items(request, reportclass.basequeryset).using(request.database)
    recs = query.count()
    total_pages = math.ceil(float(recs) / request.pagesize)
    if page > total_pages:
      page = total_pages
    if page < 1:
      page = 1
    prefs = request.user.getPreference(reportclass.getKey())
    query = reportclass._apply_sort(request, query, prefs)

    yield '{"total":%d,\n' % total_pages
    yield '"page":%d,\n' % page
    yield '"records":%d,\n' % recs
    yield '"rows":[\n'
    cnt = (page - 1) * request.pagesize + 1
    first = True

    # GridReport
    fields = [ i.field_name for i in reportclass.rows if i.field_name ]
    for i in hasattr(reportclass, 'query') and reportclass.query(request, query) or query[cnt - 1:cnt + request.pagesize].values(*fields):
      if first:
        r = [ '{' ]
        first = False
      else:
        r = [ ',\n{' ]
      first2 = True
      for f in reportclass.rows:
        if not f.name:
          continue
        if isinstance(i[f.field_name], str) or isinstance(i[f.field_name], (list, tuple)):
          s = json.dumps(i[f.field_name])
        else:
          s = '"%s"' % i[f.field_name]
        if first2:
          # if isinstance(i[f.field_name], (list,tuple)): pegging report has a tuple of strings...
          r.append('"%s":%s' % (f.name, s))
          first2 = False
        elif i[f.field_name] is not None:
          r.append(', "%s":%s' % (f.name, s))
      r.append('}')
      yield ''.join(r)
    yield '\n]}\n'


  @classmethod
  def post(reportclass, request, *args, **kwargs):
    if "csv_file" in request.FILES:
      # Note: the detection of the type of uploaded file depends on the
      # browser setting the right mime type of the file.
      if (request.FILES['csv_file'].content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
        # Uploading a spreadsheet file
        return StreamingHttpResponse(
          content_type='text/plain; charset=%s' % settings.DEFAULT_CHARSET,
          streaming_content=reportclass.parseSpreadsheetUpload(request)
          )
      else:
        # Uploading a CSV file
        return StreamingHttpResponse(
          content_type='text/plain; charset=%s' % settings.DEFAULT_CHARSET,
          streaming_content=reportclass.parseCSVupload(request)
          )
    else:
      # Saving after inline edits
      return reportclass.parseJSONupload(request)


  @classmethod
  def get(reportclass, request, *args, **kwargs):
    # Pick up the list of time buckets
    if reportclass.hasTimeBuckets:
      reportclass.getBuckets(request, args, kwargs)
      if reportclass.maxBucketLevel:
        bucketnames = Bucket.objects.order_by('-level').filter(level__lte=reportclass.maxBucketLevel).values_list('name', flat=True)
        if request.report_bucket not in bucketnames:
          # Current preference is set to a higher granularity than allowed.
          # We adjust to the highest supported level.
          request.user.horizonbuckets = bucketnames[0]
          reportclass.getBuckets(request, args, kwargs)
      else:
        bucketnames = Bucket.objects.order_by('-level').values_list('name', flat=True)
    else:
      bucketnames = None
    fmt = request.GET.get('format', None)
    if not fmt:
      # Return HTML page
      reportkey = reportclass.getKey()
      prefs = request.user.getPreference(reportkey)
      if not hasattr(reportclass, 'crosses'):
        cross_idx = None
        cross_list = None
      elif prefs and 'crosses' in prefs:
        cross_idx = ','.join([str(i) for i in prefs['crosses']])
        cross_list = reportclass._render_cross()
      else:
        cross_idx = ','.join([str(i) for i in range(len(reportclass.crosses)) if reportclass.crosses[i][1].get('visible', True)])
        cross_list = reportclass._render_cross()
      if args:
        mode = "table"
      else:
        mode = request.GET.get('mode', None)
        if mode:
          # Store the mode passed in the URL on the session to remember for the next report
          request.session['mode'] = mode
        else:
          # Pick up the mode from the session
          mode = request.session.get('mode', 'graph')
      is_popup = '_popup' in request.GET
      context = {
        'reportclass': reportclass,
        'title': (args and args[0] and _('%(title)s for %(entity)s') % {'title': force_text(reportclass.title), 'entity': force_text(args[0])}) or reportclass.title,
        'preferences': prefs,
        'reportkey': reportkey,
        'colmodel': reportclass._render_colmodel(is_popup, prefs, mode),
        'cross_idx': cross_idx,
        'cross_list': cross_list,
        'object_id': args and quote(args[0]) or None,
        'preferences': prefs,
        'page': prefs and prefs.get('page', 1) or 1,
        'sord': prefs and prefs.get('sord', request.GET.get('sord', 'asc')) or request.GET.get('sord', 'asc'),
        'sidx': prefs and prefs.get('sidx', request.GET.get('sidx', '')) or request.GET.get('sidx', ''),
        'is_popup': is_popup,
        'filters': reportclass.getQueryString(request) or (prefs and prefs.get('filter', None)),
        'args': args,
        'bucketnames': bucketnames,
        'model': reportclass.model,
        'adminsite': reportclass.adminsite,
        'hasaddperm': reportclass.editable and reportclass.model and request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, get_permission_codename('add', reportclass.model._meta))),
        'hasdeleteperm': reportclass.editable and reportclass.model and request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, get_permission_codename('delete', reportclass.model._meta))),
        'haschangeperm': reportclass.editable and reportclass.model and request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, get_permission_codename('change', reportclass.model._meta))),
        'active_tab': 'plan',
        'mode': mode,
        'actions': reportclass.actions
        }
      for k, v in reportclass.extra_context(request, *args, **kwargs).items():
        context[k] = v
      return render(request, reportclass.template, context)
    elif fmt == 'json':
      # Return JSON data to fill the grid.
      response = StreamingHttpResponse(
        content_type='application/json; charset=%s' % settings.DEFAULT_CHARSET,
        streaming_content=reportclass._generate_json_data(request, *args, **kwargs)
        )
      response['Cache-Control'] = "no-cache, no-store"
      return response
    elif fmt in ('spreadsheetlist', 'spreadsheettable', 'spreadsheet'):
      # Return an excel spreadsheet
      return reportclass._generate_spreadsheet_data(request, *args, **kwargs)
    elif fmt in ('csvlist', 'csvtable', 'csv'):
      # Return CSV data to export the data
      response = StreamingHttpResponse(
        content_type='text/csv; charset=%s' % settings.CSV_CHARSET,
        streaming_content=reportclass._generate_csv_data(request, *args, **kwargs)
        )
      response['Content-Disposition'] = 'attachment; filename=%s.csv' % iri_to_uri(reportclass.title.lower())
      response['Cache-Control'] = "no-cache, no-store"
      return response
    else:
      raise Http404('Unknown format type')


  @classmethod
  def parseJSONupload(reportclass, request):
    # Check permissions
    if not reportclass.model or not reportclass.editable:
      return HttpResponseForbidden(_('Permission denied'))
    permname = get_permission_codename('change', reportclass.model._meta)
    if not request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, permname)):
      return HttpResponseForbidden(_('Permission denied'))

    # Loop over the data records
    resp = HttpResponse()
    ok = True
    with transaction.atomic(using=request.database, savepoint=False):
      content_type_id = ContentType.objects.get_for_model(reportclass.model).pk
      for rec in json.JSONDecoder().decode(request.read().decode(request.encoding or settings.DEFAULT_CHARSET)):
        if 'delete' in rec:
          # Deleting records
          for key in rec['delete']:
            try:
              with transaction.atomic(using=request.database, savepoint=False):
                obj = reportclass.model.objects.using(request.database).get(pk=key)
                obj.delete()
                LogEntry(
                  user_id=request.user.id,
                  content_type_id=content_type_id,
                  object_id=force_text(key),
                  object_repr=force_text(key)[:200],
                  action_flag=DELETION
                ).save(using=request.database)
            except reportclass.model.DoesNotExist:
              ok = False
              resp.write(escape(_("Can't find %s" % key)))
              resp.write('<br/>')
            except Exception as e:
              ok = False
              resp.write(escape(e))
              resp.write('<br/>')
        elif 'copy' in rec:
          # Copying records
          for key in rec['copy']:
            try:
              with transaction.atomic(using=request.database, savepoint=False):
                obj = reportclass.model.objects.using(request.database).get(pk=key)
                if isinstance(reportclass.model._meta.pk, CharField):
                  # The primary key is a string
                  obj.pk = "Copy of %s" % key
                elif isinstance(reportclass.model._meta.pk, AutoField):
                  # The primary key is an auto-generated number
                  obj.pk = None
                else:
                  raise Exception(_("Can't copy %s") % reportclass.model._meta.app_label)
                obj.save(using=request.database, force_insert=True)
                LogEntry(
                  user_id=request.user.pk,
                  content_type_id=content_type_id,
                  object_id=obj.pk,
                  object_repr=force_text(obj),
                  action_flag=ADDITION,
                  change_message=_('Copied from %s.') % key
                ).save(using=request.database)
            except reportclass.model.DoesNotExist:
              ok = False
              resp.write(escape(_("Can't find %s" % key)))
              resp.write('<br/>')
            except Exception as e:
              ok = False
              resp.write(escape(e))
              resp.write('<br/>')
        else:
          # Editing records
          try:
            with transaction.atomic(using=request.database, savepoint=False):
              obj = reportclass.model.objects.using(request.database).get(pk=rec['id'])
              del rec['id']
              UploadForm = modelform_factory(
                reportclass.model,
                fields=tuple(rec.keys()),
                formfield_callback=lambda f: (isinstance(f, RelatedField) and f.formfield(using=request.database)) or f.formfield()
                )
              form = UploadForm(rec, instance=obj)
              if form.has_changed():
                obj = form.save(commit=False)
                obj.save(using=request.database)
                LogEntry(
                  user_id=request.user.pk,
                  content_type_id=content_type_id,
                  object_id=obj.pk,
                  object_repr=force_text(obj),
                  action_flag=CHANGE,
                  change_message=_('Changed %s.') % get_text_list(form.changed_data, _('and'))
                ).save(using=request.database)
          except reportclass.model.DoesNotExist:
            ok = False
            resp.write(escape(_("Can't find %s" % rec['id'])))
            resp.write('<br/>')
          except (ValidationError, ValueError):
            ok = False
            for error in form.non_field_errors():
              resp.write(escape('%s: %s' % (rec['id'], error)))
              resp.write('<br/>')
            for field in form:
              for error in field.errors:
                resp.write(escape('%s %s: %s: %s' % (obj.pk, field.name, rec[field.name], error)))
                resp.write('<br/>')
          except Exception as e:
            ok = False
            resp.write(escape(e))
            resp.write('<br/>')
    if ok:
      resp.write("OK")
    resp.status_code = ok and 200 or 500
    return resp


  @staticmethod
  def dependent_models(m, found):
    ''' An auxilary method that constructs a set of all dependent models'''
    for f in m._meta.get_fields():
      if f.is_relation and f.auto_created and f.related_model != m and f.related_model not in found:
        found.update([f.related_model])
        GridReport.dependent_models(f.related_model, found)


  @classmethod
  def erase(reportclass, request):
    # Build a list of dependencies
    deps = set([reportclass.model])
    GridReport.dependent_models(reportclass.model, deps)

    # Check the delete permissions for all related objects
    for m in deps:
      permname = get_permission_codename('delete', m._meta)
      if not request.user.has_perm('%s.%s' % (m._meta.app_label, permname)):
        return string_concat(m._meta.verbose_name, ':', _('Permission denied'))

    # Delete the data records
    cursor = connections[request.database].cursor()
    with transaction.atomic(using=request.database):
      sql_list = connections[request.database].ops.sql_flush(no_style(), [m._meta.db_table for m in deps], [] )
      for sql in sql_list:
        cursor.execute(sql)
      # Erase comments and history
      content_ids = [ContentType.objects.get_for_model(m) for m in deps]
      LogEntry.objects.filter(content_type__in=content_ids).delete()
      Comment.objects.filter(content_type__in=content_ids).delete()
      # Prepare message
      for m in deps:
        messages.add_message(
          request, messages.INFO,
          _('Erasing data from %(model)s') % {'model': force_text(m._meta.verbose_name)}
          )

    # Finished successfully
    return None


  @classmethod
  def parseCSVupload(reportclass, request):
      '''
      This method reads CSV data from a string (in memory) and creates or updates
      the database records.
      The data must follow the following format:
        - the first row contains a header, listing all field names
        - a first character # marks a comment line
        - empty rows are skipped
      '''
      # Check permissions
      if not reportclass.model:
        yield force_text(_('Invalid upload request')) + '\n '
        return
      permname = get_permission_codename('add', reportclass.model._meta)
      if not reportclass.editable or not request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, permname)):
        yield force_text(_('Permission denied')) + '\n '
        return

      # Choose the right delimiter and language
      delimiter = get_format('DECIMAL_SEPARATOR', request.LANGUAGE_CODE, True) == ',' and ';' or ','
      if translation.get_language() != request.LANGUAGE_CODE:
        translation.activate(request.LANGUAGE_CODE)

      # Init
      headers = []
      rownumber = 0
      changed = 0
      added = 0
      content_type_id = ContentType.objects.get_for_model(reportclass.model).pk

      # Handle the complete upload as a single database transaction
      with transaction.atomic(using=request.database):

        # Erase all records and related tables
        errors = False
        if 'erase' in request.POST:
          returnvalue = reportclass.erase(request)
          if returnvalue:
            yield returnvalue + '\n '
            errors = True

        # Loop through the data records
        has_pk_field = False
        for row in EncodedCSVReader(request.FILES['csv_file'], delimiter=delimiter):
          rownumber += 1

          ### Case 1: The first line is read as a header line
          if rownumber == 1:

            for col in row:
              col = col.strip().strip('#').lower()
              if col == "":
                headers.append(False)
                continue
              ok = False
              for i in reportclass.model._meta.fields:
                if col == i.name.lower() or col == i.verbose_name.lower():
                  if i.editable is True:
                    headers.append(i)
                  else:
                    headers.append(False)
                  ok = True
                  break
              if not ok:
                errors = True
                yield force_text(_('Incorrect field %(column)s') % {'column': col}) + '\n '
              if col == reportclass.model._meta.pk.name.lower() or \
                 col == reportclass.model._meta.pk.verbose_name.lower():
                has_pk_field = True
            if not has_pk_field and not isinstance(reportclass.model._meta.pk, AutoField):
              # The primary key is not an auto-generated id and it is not mapped in the input...
              errors = True
              yield force_text(_('Missing primary key field %(key)s') % {'key': reportclass.model._meta.pk.name}) + '\n '
            # Abort when there are errors
            if errors:
              break

            # Create a form class that will be used to validate the data
            UploadForm = modelform_factory(
              reportclass.model,
              fields=tuple([i.name for i in headers if isinstance(i, Field)]),
              formfield_callback=lambda f: (isinstance(f, RelatedField) and f.formfield(using=request.database, localize=True)) or f.formfield(localize=True)
              )

          ### Case 2: Skip empty rows and comments rows
          elif len(row) == 0 or row[0].startswith('#'):
            continue

          ### Case 3: Process a data row
          else:
            try:
              # Step 1: Build a dictionary with all data fields
              d = {}
              colnum = 0
              for col in row:
                # More fields in data row than headers. Move on to the next row.
                if colnum >= len(headers):
                  break
                if isinstance(headers[colnum], Field):
                  d[headers[colnum].name] = col
                colnum += 1

              # Step 2: Fill the form with data, either updating an existing
              # instance or creating a new one.
              if has_pk_field:
                # A primary key is part of the input fields
                try:
                  # Try to find an existing record with the same primary key
                  it = reportclass.model.objects.using(request.database).get(pk=d[reportclass.model._meta.pk.name])
                  form = UploadForm(d, instance=it)
                except reportclass.model.DoesNotExist:
                  form = UploadForm(d)
                  it = None
              else:
                # No primary key required for this model
                form = UploadForm(d)
                it = None

              # Step 3: Validate the data and save to the database
              if form.has_changed():
                try:
                  with transaction.atomic(using=request.database):
                    obj = form.save(commit=False)
                    obj.save(using=request.database)
                    LogEntry(
                      user_id=request.user.pk,
                      content_type_id=content_type_id,
                      object_id=obj.pk,
                      object_repr=force_text(obj),
                      action_flag=it and CHANGE or ADDITION,
                      change_message=_('Changed %s.') % get_text_list(form.changed_data, _('and'))
                    ).save(using=request.database)
                    if it:
                      changed += 1
                    else:
                      added += 1
                except Exception as e:
                  # Validation fails
                  for error in form.non_field_errors():
                    yield force_text(
                      _('Row %(rownum)s: %(message)s') % {
                        'rownum': rownumber, 'message': error
                      }) + '\n '
                  for field in form:
                    for error in field.errors:
                      yield force_text(
                        _('Row %(rownum)s field %(field)s: %(data)s: %(message)s') % {
                          'rownum': rownumber, 'data': d[field.name],
                          'field': field.name, 'message': error
                        }) + '\n '
            except Exception as e:
              yield force_text(_("Exception during upload: %(message)s") % {'message': e}) + '\n '

      # Report all failed records
      yield force_text(
          _('Uploaded data successfully: changed %(changed)d and added %(added)d records') % {'changed': changed, 'added': added}
          ) + '\n '


  @classmethod
  def parseSpreadsheetUpload(reportclass, request):
    '''
    This method reads a spreadsheet file (in memory) and creates or updates
    the database records.
    The data must follow the following format:
      - only the first tab in the spreadsheet is read
      - the first row contains a header, listing all field names
      - a first character # marks a comment line
      - empty rows are skipped
    '''
    # Check permissions
    if not reportclass.model:
      yield force_text(_('Invalid upload request')) + '\n '
      return
    permname = get_permission_codename('add', reportclass.model._meta)
    if not reportclass.editable or not request.user.has_perm('%s.%s' % (reportclass.model._meta.app_label, permname)):
      yield force_text(_('Permission denied')) + '\n '
      return

    # Choose the right language
    if translation.get_language() != request.LANGUAGE_CODE:
      translation.activate(request.LANGUAGE_CODE)

    # Init
    headers = []
    rownumber = 0
    changed = 0
    added = 0
    content_type_id = ContentType.objects.get_for_model(reportclass.model).pk

    # Handle the complete upload as a single database transaction
    with transaction.atomic(using=request.database):

      # Erase all records and related tables
      errors = False
      if 'erase' in request.POST:
        returnvalue = reportclass.erase(request)
        if returnvalue:
          errors = True
          yield returnvalue + '\n '

      # Loop through the data records
      wb = load_workbook(filename=request.FILES['csv_file'], use_iterators=True, data_only=True)
      ws = wb.worksheets[0]
      has_pk_field = False
      for row in ws.iter_rows():
        rownumber += 1

        ### Case 1: The first line is read as a header line
        if rownumber == 1:
          for col in row:
            col = str(col.value).strip().strip('#').lower()
            if col == "":
              headers.append(False)
              continue
            ok = False
            for i in reportclass.model._meta.fields:
              if col.replace(' ', '') == i.name.lower() or col == i.verbose_name.lower():
                if i.editable is True:
                  headers.append(i)
                else:
                  headers.append(False)
                ok = True
                break
            if not ok:
              errors = True
              yield force_text(_('Incorrect field %(column)s') % {'column': col}) + '\n '
            if col == reportclass.model._meta.pk.name.lower() or \
               col == reportclass.model._meta.pk.verbose_name.lower():
              has_pk_field = True
          if not has_pk_field and not isinstance(reportclass.model._meta.pk, AutoField):
            # The primary key is not an auto-generated id and it is not mapped in the input...
            errors = True
            yield force_text(_('Missing primary key field %(key)s') % {'key': reportclass.model._meta.pk.name}) + '\n '
          # Abort when there are errors
          if errors > 0:
            break

          # Create a form class that will be used to validate the data
          UploadForm = modelform_factory(
            reportclass.model,
            fields=tuple([i.name for i in headers if isinstance(i, Field)]),
            formfield_callback=lambda f: (isinstance(f, RelatedField) and f.formfield(using=request.database, localize=True)) or f.formfield(localize=True)
            )

        ### Case 2: Skip empty rows and comments rows
        elif len(row) == 0 or (isinstance(row[0].value, six.string_types) and row[0].value.startswith('#')):
          continue

        ### Case 3: Process a data row
        else:
          try:
            # Step 1: Build a dictionary with all data fields
            d = {}
            colnum = 0
            for col in row:
              # More fields in data row than headers. Move on to the next row.
              if colnum >= len(headers):
                break
              if isinstance(headers[colnum], Field):
                data = col.value
                if isinstance(headers[colnum], (IntegerField, AutoField)):
                  if isinstance(data, numericTypes):
                    data = int(data)
                d[headers[colnum].name] = data
              colnum += 1

            # Step 2: Fill the form with data, either updating an existing
            # instance or creating a new one.
            if has_pk_field:
              # A primary key is part of the input fields
              try:
                # Try to find an existing record with the same primary key
                it = reportclass.model.objects.using(request.database).get(pk=d[reportclass.model._meta.pk.name])
                form = UploadForm(d, instance=it)
              except reportclass.model.DoesNotExist:
                form = UploadForm(d)
                it = None
            else:
              # No primary key required for this model
              form = UploadForm(d)
              it = None

            # Step 3: Validate the data and save to the database
            if form.has_changed():
              try:
                with transaction.atomic(using=request.database):
                  obj = form.save(commit=False)
                  obj.save(using=request.database)
                  LogEntry(
                    user_id=request.user.pk,
                    content_type_id=content_type_id,
                    object_id=obj.pk,
                    object_repr=force_text(obj),
                    action_flag=it and CHANGE or ADDITION,
                    change_message=_('Changed %s.') % get_text_list(form.changed_data, _('and'))
                  ).save(using=request.database)
                  if it:
                    changed += 1
                  else:
                    added += 1
              except Exception as e:
                # Validation fails
                for error in form.non_field_errors():
                  yield force_text(
                    _('Row %(rownum)s: %(message)s') % {
                      'rownum': rownumber, 'message': error
                    }) + '\n '
                for field in form:
                  for error in field.errors:
                    yield force_text(
                      _('Row %(rownum)s field %(field)s: %(data)s: %(message)s') % {
                        'rownum': rownumber, 'data': d[field.name],
                        'field': field.name, 'message': error
                      }) + '\n '
          except Exception as e:
            yield force_text(_("Exception during upload: %(message)s") % {'message': e}) + '\n '

    # Report all failed records
    yield force_text(
      _('Uploaded data successfully: changed %(changed)d and added %(added)d records') % {'changed': changed, 'added': added}
      ) + '\n '


  @classmethod
  def _getRowByName(reportclass, name):
    if not hasattr(reportclass, '_rowsByName'):
      reportclass._rowsByName = {}
      for i in reportclass.rows:
        reportclass._rowsByName[i.name] = i
        if i.field_name != i.name:
          reportclass._rowsByName[i.field_name] = i
    return reportclass._rowsByName[name]


  _filter_map_jqgrid_django = {
      # jqgrid op: (django_lookup, use_exclude)
      'ne': ('%(field)s__exact', True),
      'bn': ('%(field)s__startswith', True),
      'en': ('%(field)s__endswith', True),
      'nc': ('%(field)s__contains', True),
      'ni': ('%(field)s__in', True),
      'in': ('%(field)s__in', False),
      'eq': ('%(field)s__exact', False),
      'bw': ('%(field)s__startswith', False),
      'gt': ('%(field)s__gt', False),
      'ge': ('%(field)s__gte', False),
      'lt': ('%(field)s__lt', False),
      'le': ('%(field)s__lte', False),
      'ew': ('%(field)s__endswith', False),
      'cn': ('%(field)s__contains', False)
  }


  _filter_map_django_jqgrid = {
      # django lookup: jqgrid op
      'in': 'in',
      'exact': 'eq',
      'startswith': 'bw',
      'gt': 'gt',
      'gte': 'ge',
      'lt': 'lt',
      'lte': 'le',
      'endswith': 'ew',
      'contains': 'cn',
  }


  @classmethod
  def getQueryString(reportclass, request):
    # Django-style filtering (which uses URL parameters) are converted to a jqgrid filter expression
    filtered = False
    filters = ['{"groupOp":"AND","rules":[']
    for i, j in request.GET.items():
      for r in reportclass.rows:
        if r.field_name and i.startswith(r.field_name):
          operator = (i == r.field_name) and 'exact' or i[i.rfind('_') + 1:]
          try:
            filters.append(
              '{"field":"%s","op":"%s","data":"%s"},' % (
              r.field_name, reportclass._filter_map_django_jqgrid[operator], unquote(j).replace('"', '\\"')
              ))
            filtered = True
          except:
            pass  # Ignore invalid operators
    if not filtered:
      return None
    filters.append(']}')
    return ''.join(filters)


  @classmethod
  def _get_q_filter(reportclass, filterdata):
    q_filters = []
    for rule in filterdata['rules']:
      try:
        op, field, data = rule['op'], rule['field'], rule['data']
        filter_fmt, exclude = reportclass._filter_map_jqgrid_django[op]
        reportrow = reportclass._getRowByName(field)
        if data == '' and not isinstance(reportrow, (GridFieldText, GridFieldChoice)):
          # Filter value specified, which makes the filter invalid
          continue
        filter_str = smart_str(filter_fmt % {'field': reportrow.field_name})
        if filter_fmt.endswith('__in'):
          filter_kwargs = {filter_str: data.split(',')}
        else:
          filter_kwargs = {filter_str: smart_str(data)}
        if exclude:
          q_filters.append(~models.Q(**filter_kwargs))
        else:
          q_filters.append(models.Q(**filter_kwargs))
      except:
        pass  # Silently ignore invalid filters
    if 'groups' in filterdata:
      for group in filterdata['groups']:
        try:
          z = reportclass._get_q_filter(group)
          if z:
            q_filters.append(z)
        except:
          pass  # Silently ignore invalid groups
    if len(q_filters) == 0:
      return None
    elif filterdata['groupOp'].upper() == 'OR':
      return functools.reduce(operator.ior, q_filters)
    else:
      return functools.reduce(operator.iand, q_filters)


  @classmethod
  def filter_items(reportclass, request, items, plus_django_style=True):
    # Jqgrid-style advanced filtering
    filters = None
    _filters = request.GET.get('filters')
    if _filters:
      # Validate complex search JSON data
      try:
        filters = _filters and json.loads(_filters)
      except ValueError:
        filters = None

    # Single field searching, which is currently not used
    if request.GET.get('_search') == 'true' and not filters:
      field = request.GET.get('searchField')
      op = request.GET.get('searchOper')
      data = request.GET.get('searchString')
      if all([field, op, data]):
        filters = {
            'groupOp': 'AND',
            'rules': [{ 'op': op, 'field': field, 'data': data }]
        }

    if filters:
      z = reportclass._get_q_filter(filters)
      if z:
        return items.filter(z)
      else:
        return items

    # Django-style filtering, using URL parameters
    if plus_django_style:
      for i, j in request.GET.items():
        for r in reportclass.rows:
          if r.name and i.startswith(r.field_name):
            try:
              items = items.filter(**{i: unquote(j)})
            except:
              pass  # silently ignore invalid filters
    return items


class GridPivot(GridReport):

  # Cross definitions.
  # Possible attributes for a cross field are:
  #   - title:
  #     Name of the cross that is displayed to the user.
  #     It defaults to the name of the field.
  #   - editable:
  #     True when the field is editable in the page.
  #     The default value is false.
  crosses = ()

  template = 'admin/base_site_gridpivot.html'

  hasTimeBuckets = True

  editable = False

  multiselect = False


  @classmethod
  def _render_cross(cls):
    result = []
    for i in cls.crosses:
      result.append(
        "{name:'%s',editable:%s}"
        % (title('title' in i[1] and i[1]['title'] or ''), i[1].get('editable', False) and 'true' or 'false')
        )
    return ',\n'.join(result)


  @classmethod
  def _render_colmodel(cls, is_popup=False, prefs=None, mode="graph"):
    if prefs and 'rows' in prefs:
      rows = prefs['rows']
    else:
      rows = [ (i, cls.rows[i].hidden, cls.rows[i].width) for i in range(len(cls.rows)) ]

    result = []
    if is_popup:
      result.append("{name:'select',label:gettext('Select'),width:75,align:'center',sortable:false,search:false,fixed:true}")
    count = -1
    for (index, hidden, width) in rows:
      count += 1
      result.append("{%s,width:%s,counter:%d,frozen:true%s,hidden:%s,searchoptions:{searchhidden: true},fixed:true}" % (
        cls.rows[index], width, index,
        is_popup and ',popup:true' or '',
        hidden and 'true' or 'false'
        ))
    if mode == "graph":
      result.append(
        "{name:'graph',index:'graph',editable:false,label:' ',title:false,"
        "sortable:false,formatter:'graph',searchoptions:{searchhidden: true},fixed:false}"
        )
    else:
      result.append(
        "{name:'columns',label:' ',sortable:false,width:150,align:'left',"
        "formatter:grid.pivotcolumns,search:false,frozen:true,title:false }"
        )
    return ',\n'.join(result)


  @classmethod
  def _apply_sort(reportclass, request, query, prefs=None):
    '''
    Applies a sort to the query.
    '''
    asc = True
    sort = None
    if 'sidx' in request.GET:
      sort = request.GET['sidx']
      if 'sord' in request.GET and request.GET['sord'] == 'desc':
        asc = False
    if not sort:
      if prefs and 'sidx' in prefs:
        sort = prefs['sidx']
        if 'sord' in prefs and prefs['sord'] == 'desc':
          asc = False
      if not sort and reportclass.default_sort:
        sort = reportclass.rows[reportclass.default_sort[0]].name
        if reportclass.default_sort[1] == 'desc':
          asc = False
      else:
        # No sorting
        return query
    if sort:
      # Validate the field does exist.
      for i in reportclass.rows:
        if i.name == sort and i.search:
          return query.order_by(asc and i.field_name or ('-%s' % i.field_name))
    # Sorting by a non-existent field name: ignore the filter
      return query


  @classmethod
  def _apply_sort_index(reportclass, request, prefs=None):
    '''
    Returns the index of the column to sort on.
    '''
    if 'sidx' in request.GET:
      sort = request.GET['sidx']
    elif prefs and 'sidx' in prefs:
      sort = prefs['sidx']
    else:
      sort = reportclass.rows[0].name
    idx = 1
    for i in reportclass.rows:
      if i.name == sort:
        if 'sord' in request.GET and request.GET['sord'] == 'desc':
          return idx > 1 and "%d desc, 1 asc" % idx or "1 desc"
        elif prefs and 'sord' in prefs and prefs['sord'] == 'desc':
          return idx > 1 and "%d desc, 1 asc" % idx or "1 desc"
        else:
          return idx > 1 and "%d asc, 1 asc" % idx or "1 asc"
      else:
        idx += 1
    return "1 asc"


  @classmethod
  def _generate_json_data(reportclass, request, *args, **kwargs):
    # Prepare the query
    if args and args[0]:
      page = 1
      recs = 1
      total_pages = 1
      query = reportclass.query(request, reportclass.basequeryset.filter(pk__exact=args[0]).using(request.database), sortsql="1 asc")
    else:
      page = 'page' in request.GET and int(request.GET['page']) or 1
      if isinstance(reportclass.basequeryset, collections.Callable):
        recs = reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database).count()
      else:
        recs = reportclass.filter_items(request, reportclass.basequeryset).using(request.database).count()
      total_pages = math.ceil(float(recs) / request.pagesize)
      if page > total_pages:
        page = total_pages
      if page < 1:
        page = 1
      cnt = (page - 1) * request.pagesize + 1
      prefs = request.user.getPreference(reportclass.getKey())
      if isinstance(reportclass.basequeryset, collections.Callable):
        query = reportclass.query(
          request,
          reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False), prefs).using(request.database)[cnt - 1:cnt + request.pagesize],
          sortsql=reportclass._apply_sort_index(request, prefs)
          )
      else:
        query = reportclass.query(
          request,
          reportclass._apply_sort(request, reportclass.filter_items(request, reportclass.basequeryset), prefs).using(request.database)[cnt - 1:cnt + request.pagesize],
          sortsql=reportclass._apply_sort_index(request, prefs)
          )

    # Generate header of the output
    yield '{"total":%d,\n' % total_pages
    yield '"page":%d,\n' % page
    yield '"records":%d,\n' % recs
    yield '"rows":[\n'

    # Generate output
    currentkey = None
    r = []
    for i in query:
      # We use the first field in the output to recognize new rows.
      if currentkey != i[reportclass.rows[0].name]:
        # New line
        if currentkey:
          yield ''.join(r)
          r = [ '},\n{' ]
        else:
          r = [ '{' ]
        currentkey = i[reportclass.rows[0].name]
        first2 = True
        for f in reportclass.rows:
          try:
            s = isinstance(i[f.name], str) and escape(i[f.name].encode(settings.DEFAULT_CHARSET, "ignore")) or i[f.name]
            if first2:
              r.append('"%s":"%s"' % (f.name, s))
              first2 = False
            elif i[f.name] is not None:
              r.append(', "%s":"%s"' % (f.name, s))
          except:
            pass
      r.append(', "%s":[' % i['bucket'])
      first2 = True
      for f in reportclass.crosses:
        if i[f[0]] == None:
          if first2:
            r.append('null')
            first2 = False
          else:
            r.append(',null')
        else:
          if first2:
            r.append('%s' % i[f[0]])
            first2 = False
          else:
            r.append(',%s' % i[f[0]])
      r.append(']')
    r.append('}')
    r.append('\n]}\n')
    yield ''.join(r)


  @classmethod
  def _generate_csv_data(reportclass, request, *args, **kwargs):
    sf = StringIO()
    decimal_separator = get_format('DECIMAL_SEPARATOR', request.LANGUAGE_CODE, True)
    if decimal_separator == ',':
      writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=';')
    else:
      writer = csv.writer(sf, quoting=csv.QUOTE_NONNUMERIC, delimiter=',')
    if translation.get_language() != request.LANGUAGE_CODE:
      translation.activate(request.LANGUAGE_CODE)
    listformat = (request.GET.get('format', 'csvlist') == 'csvlist')

    # Prepare the query
    prefs = request.user.getPreference(reportclass.getKey())
    if args and args[0]:
      query = reportclass.query(request, reportclass.basequeryset.filter(pk__exact=args[0]).using(request.database), sortsql="1 asc")
    elif isinstance(reportclass.basequeryset, collections.Callable):
      query = reportclass.query(request, reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database), sortsql=reportclass._apply_sort_index(request, prefs))
    else:
      query = reportclass.query(request, reportclass.filter_items(request, reportclass.basequeryset).using(request.database), sortsql=reportclass._apply_sort_index(request, prefs))

    # Write a Unicode Byte Order Mark header, aka BOM (Excel needs it to open UTF-8 file properly)
    encoding = settings.CSV_CHARSET
    yield getBOM(encoding)

    # Pick up the preferences
    if prefs and 'rows' in prefs:
      myrows = [
        reportclass.rows[f[0]]
        for f in prefs['rows']
        if not f[1] and f[0] < len(reportclass.rows)
        ]
    else:
      myrows = [ f for f in reportclass.rows if f.name and not f.hidden ]
    if prefs and 'crosses' in prefs:
      mycrosses = [ reportclass.crosses[f] for f in prefs['crosses'] if f < len(reportclass.crosses) and reportclass.crosses[f][1].get('visible', True) ]
    else:
      mycrosses = [ f for f in reportclass.crosses if f[1].get('visible', True) ]

    # Write a header row
    fields = [
      force_text(f.title, encoding=encoding, errors='ignore').title()
      for f in myrows if f.name
      ]
    if listformat:
      fields.extend([ capfirst(force_text(_('bucket'), encoding=encoding, errors='ignore')) ])
      fields.extend([ capfirst(force_text(_(f[1].get('title', _(f[0]))), encoding=encoding, errors='ignore')) for f in mycrosses ])
    else:
      fields.extend([ capfirst(force_text(_('data field'), encoding=encoding, errors='ignore')) ])
      fields.extend([ force_text(b['name'], encoding=encoding, errors='ignore') for b in request.report_bucketlist])
    writer.writerow(fields)
    yield sf.getvalue()

    # Write the report content
    if listformat:
      for row in query:
        # Clear the return string buffer
        sf.seek(0)
        sf.truncate(0)
        # Data for rows
        if hasattr(row, "__getitem__"):
          fields = [
            force_text(row[f.name], encoding=encoding, errors='ignore') if row[f.name] is not None else ''
            for f in myrows if f.name
            ]
          fields.extend([ force_text(row['bucket'], encoding=encoding, errors='ignore') ])
          fields.extend([
            force_text(_localize(row[f[0]], decimal_separator), encoding=encoding, errors='ignore') if row[f[0]] is not None else ''
            for f in mycrosses
            ])
        else:
          fields = [
            force_text(getattr(row, f.name), encoding=encoding, errors='ignore') if getattr(row, f.name) is not None else ''
            for f in myrows if f.name
            ]
          fields.extend([ force_text(getattr(row, 'bucket'), encoding=encoding, errors='ignore') ])
          fields.extend([
            force_text(_localize(getattr(row, f[0]), decimal_separator), encoding=encoding, errors='ignore') if getattr(row, f[0]) is not None else ''
            for f in mycrosses
            ])
        # Return string
        writer.writerow(fields)
        yield sf.getvalue()
    else:
      currentkey = None
      for row in query:
        # We use the first field in the output to recognize new rows.
        if not currentkey:
          currentkey = row[reportclass.rows[0].name]
          row_of_buckets = [ row ]
        elif currentkey == row[reportclass.rows[0].name]:
          row_of_buckets.append(row)
        else:
          # Write an entity
          for cross in mycrosses:
            # Clear the return string buffer
            sf.seek(0)
            sf.truncate(0)
            fields = [
              force_text(row_of_buckets[0][s.name], encoding=encoding, errors='ignore')
              for s in myrows  if s.name
              ]
            fields.extend([
              force_text('title' in cross[1] and capfirst(_(cross[1]['title'])) or capfirst(_(cross[0])), encoding=encoding, errors='ignore')
              ])
            fields.extend([
              force_text(_localize(bucket[cross[0]], decimal_separator), encoding=encoding, errors='ignore')
              for bucket in row_of_buckets
              ])
            # Return string
            writer.writerow(fields)
            yield sf.getvalue()
          currentkey = row[reportclass.rows[0].name]
          row_of_buckets = [row]
      # Write the last entity
      for cross in mycrosses:
        # Clear the return string buffer
        sf.seek(0)
        sf.truncate(0)
        fields = [
          force_text(row_of_buckets[0][s.name], encoding=encoding, errors='ignore')
          for s in myrows if s.name
          ]
        fields.extend([ force_text('title' in cross[1] and capfirst(_(cross[1]['title'])) or capfirst(_(cross[0])), encoding=encoding, errors='ignore') ])
        fields.extend([
          force_text(_localize(bucket[cross[0]], decimal_separator), encoding=encoding, errors='ignore')
          for bucket in row_of_buckets
          ])
        # Return string
        writer.writerow(fields)
        yield sf.getvalue()


  @classmethod
  def _generate_spreadsheet_data(reportclass, request, *args, **kwargs):
    # Create a workbook
    wb = Workbook(optimized_write=True)
    ws = wb.create_sheet(title=force_text(reportclass.model._meta.verbose_name))

    # Prepare the query
    prefs = request.user.getPreference(reportclass.getKey())
    listformat = (request.GET.get('format', 'spreadsheetlist') == 'spreadsheetlist')
    if args and args[0]:
      query = reportclass.query(request, reportclass.basequeryset.filter(pk__exact=args[0]).using(request.database), sortsql="1 asc")
    elif isinstance(reportclass.basequeryset, collections.Callable):
      query = reportclass.query(request, reportclass.filter_items(request, reportclass.basequeryset(request, args, kwargs), False).using(request.database), sortsql=reportclass._apply_sort_index(request, prefs))
    else:
      query = reportclass.query(request, reportclass.filter_items(request, reportclass.basequeryset).using(request.database), sortsql=reportclass._apply_sort_index(request, prefs))

    # Pick up the preferences
    if prefs and 'rows' in prefs:
      myrows = [
        reportclass.rows[f[0]]
        for f in prefs['rows']
        if not f[1] and f[0] < len(reportclass.rows)
        ]
    else:
      myrows = [ f for f in reportclass.rows if f.name and not f.hidden ]
    if prefs and 'crosses' in prefs:
      mycrosses = [ reportclass.crosses[f] for f in prefs['crosses'] if f < len(reportclass.crosses) and reportclass.crosses[f][1].get('visible', True) ]
    else:
      mycrosses = [ f for f in reportclass.crosses if f[1].get('visible', True) ]

    # Write a header row
    fields = [ force_text(f.title).title() for f in myrows if f.name ]
    if listformat:
      fields.extend([ capfirst(force_text(_('bucket'))) ])
      fields.extend([ capfirst(_(f[1].get('title', _(f[0])))) for f in mycrosses ])
    else:
      fields.extend( [capfirst(_('data field'))])
      fields.extend([ str(b['name']) for b in request.report_bucketlist])
    ws.append(fields)

    # Write the report content
    if listformat:
      for row in query:
        # Append a row
        if hasattr(row, "__getitem__"):
          fields = [ _getCellValue(row[f.name]) for f in myrows if f.name ]
          fields.extend([ _getCellValue(row['bucket']) ])
          fields.extend([ _getCellValue(row[f[0]]) for f in mycrosses ])
        else:
          fields = [ _getCellValue(getattr(row, f.name)) for f in myrows if f.name ]
          fields.extend([ _getCellValue(getattr(row, 'bucket')) ])
          fields.extend([ _getCellValue(getattr(row, f[0])) for f in mycrosses ])
        ws.append(fields)
    else:
      currentkey = None
      row_of_buckets = None
      for row in query:
        # We use the first field in the output to recognize new rows.
        if not currentkey:
          currentkey = row[reportclass.rows[0].name]
          row_of_buckets = [ row ]
        elif currentkey == row[reportclass.rows[0].name]:
          row_of_buckets.append(row)
        else:
          # Write a row
          for cross in mycrosses:
            if 'visible' in cross[1] and not cross[1]['visible']:
              continue
            fields = [ _getCellValue(row_of_buckets[0][s.name]) for s in myrows if s.name ]
            fields.extend([ _getCellValue(('title' in cross[1] and capfirst(_(cross[1]['title'])) or capfirst(_(cross[0])))) ])
            fields.extend([ _getCellValue(bucket[cross[0]]) for bucket in row_of_buckets ])
            ws.append(fields)
          currentkey = row[reportclass.rows[0].name]
          row_of_buckets = [row]
      # Write the last row
      if row_of_buckets:
        for cross in mycrosses:
          if 'visible' in cross[1] and not cross[1]['visible']:
            continue
          fields = [ _getCellValue(row_of_buckets[0][s.name]) for s in myrows if s.name ]
          fields.extend([ _getCellValue(('title' in cross[1] and capfirst(_(cross[1]['title'])) or capfirst(_(cross[0])))) ])
          fields.extend([ _getCellValue(bucket[cross[0]]) for bucket in row_of_buckets ])
          ws.append(fields)

    # Write the spreadsheet from memory to a string and then to a HTTP response
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      content=output.getvalue()
      )
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % reportclass.model._meta.model_name
    response['Cache-Control'] = "no-cache, no-store"
    return response


numericTypes = (Decimal, float) + six.integer_types


def _localize(value, decimal_separator):
  '''
  Localize numbers.
  Dates are always represented as YYYY-MM-DD hh:mm:ss since this is
  a format that is understood uniformly across different regions in the
  world.
  '''
  if isinstance(value, collections.Callable):
    value = value()
  if isinstance(value, numericTypes):
    return decimal_separator == "," and six.text_type(value).replace(".", ",") or six.text_type(value)
  elif isinstance(value, (list, tuple) ):
    return "|".join([ str(_localize(i, decimal_separator)) for i in value ])
  else:
    return value


def _getCellValue(data):
  if data is None:
    return ''
  if isinstance(data, numericTypes):
    return data
  return str(data)


def exportWorkbook(request):
  # Create a workbook
  wb = Workbook(optimized_write=True)

  # Loop over all selected entity types
  ok = False
  cursor = connections[request.database].cursor()
  for entity_name in request.POST.getlist('entities'):
    try:
      # Initialize
      (app_label, model_label) = entity_name.split('.')
      model = apps.get_model(app_label, model_label)
      # Verify access rights
      permname = get_permission_codename('change', model._meta)
      if not request.user.has_perm("%s.%s" % (app_label, permname)):
        continue
      # Never export some special administrative models
      if model in EXCLUDE_FROM_BULK_OPERATIONS:
        continue
      # Build a list of fields
      fields = []
      header = []
      source = False
      lastmodified = False
      for i in model._meta.fields:
        if i.name in ['lft', 'rght', 'lvl']:
          continue  # Skip some fields of HierarchyModel
        elif i.name == 'source':
          source = True  # Put the source field at the end
        elif i.name == 'lastmodified':
          lastmodified = True  # Put the last-modified field at the very end
        else:
          fields.append(connections[request.database].ops.quote_name(i.column))
          header.append(force_text(i.verbose_name))
      if source:
        fields.append("source")
        header.append(force_text(_("source")))
      if lastmodified:
        fields.append("lastmodified")
        header.append(force_text(_("last modified")))
      # Create sheet
      ok = True
      ws = wb.create_sheet(title=force_text(model._meta.verbose_name))
      # Write a header row
      ws.append(header)
      # Loop over all records
      if issubclass(model, HierarchyModel):
        model.rebuildHierarchy(database=request.database)
        cursor.execute(
          "SELECT %s FROM %s ORDER BY lvl, 1" %
          (",".join(fields), connections[request.database].ops.quote_name(model._meta.db_table))
          )
      else:
        cursor.execute(
          "SELECT %s FROM %s ORDER BY 1" %
          (",".join(fields), connections[request.database].ops.quote_name(model._meta.db_table))
          )
      for rec in cursor.fetchall():
        ws.append([ _getCellValue(f) for f in rec ])
    except:
      pass  # Silently ignore the error and move on to the next entity.

  # Not a single entity to export
  if not ok:
    raise Exception(_("Nothing to export"))

  # Write the excel from memory to a string and then to a HTTP response
  output = BytesIO()
  wb.save(output)
  response = HttpResponse(
    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    content=output.getvalue()
    )
  response['Content-Disposition'] = 'attachment; filename=frepple.xlsx'
  response['Cache-Control'] = "no-cache, no-store"
  return response


def importWorkbook(request):
  '''
  This method reads a spreadsheet in Office Open XML format (typically with
  the extension .xlsx or .ods).
  Each entity has a tab in the spreadsheet, and the first row contains
  the fields names.
  '''
  # Build a list of all contenttypes
  all_models = [ (ct.model_class(), ct.pk) for ct in ContentType.objects.all() if ct.model_class() ]
  with transaction.atomic(using=request.database):
    # Find all models in the workbook
    wb = load_workbook(filename=request.FILES['spreadsheet'], use_iterators=True, data_only=True)
    models = []
    for ws_name in wb.get_sheet_names():
      # Find the model
      model = None
      contenttype_id = None
      for m, ct in all_models:
        if ws_name.lower() in (m._meta.model_name.lower(), m._meta.verbose_name.lower(), m._meta.verbose_name_plural.lower()):
          model = m
          contenttype_id = ct
          break
      if not model or model in EXCLUDE_FROM_BULK_OPERATIONS:
        yield force_text(_("Ignoring data in worksheet: %s") % ws_name) + '\n'
      elif not request.user.has_perm('%s.%s' % (model._meta.app_label, get_permission_codename('add', model._meta))):
        # Check permissions
        yield force_text(_("You don't permissions to add: %s") % ws_name) + '\n'
      else:
        deps = set([model])
        GridReport.dependent_models(model, deps)
        models.append( (ws_name, model, contenttype_id, deps) )

    # Sort the list of models, based on dependencies between models
    cnt = len(models)
    ok = False
    while not ok:
      ok = True
      for i in range(cnt):
        for j in range(i + 1, cnt):
          if models[i][1] in models[j][3]:
            # A subsequent model i depends on model i. The list ordering is
            # thus not ok yet. We move this element to the end of the list.
            models.append(models.pop(i))
            ok = False

    # Process all rows in each worksheet
    for ws_name, model, contenttype_id, dependencies in models:
      yield force_text(_("Processing data in worksheet: %s") % ws_name) + '\n'
      ws = wb.get_sheet_by_name(name=ws_name)
      rownum = 0
      has_pk_field = False
      headers = []
      uploadform = None
      changed = 0
      added = 0
      numerrors = 0
      for row in ws.iter_rows():
        with transaction.atomic(using=request.database):
          rownum += 1
          if rownum == 1:
            # Process the header row with the field names
            header_ok = True
            for cell in row:
              ok = False
              value = cell.value
              if not value:
                headers.append(False)
                continue
              else:
                value = value.lower()
              for i in model._meta.fields:
                if value == i.name.lower() or value == i.verbose_name.lower():
                  if i.editable is True:
                    headers.append(i)
                  else:
                    headers.append(False)
                  ok = True
                  break
              if not ok:
                header_ok = False
                yield force_text(string_concat(
                  model._meta.verbose_name, ': ', _('Incorrect field %(column)s') % {'column': value}
                  )) + '\n'
                numerrors += 1
              if value == model._meta.pk.name.lower() \
                or value == model._meta.pk.verbose_name.lower():
                  has_pk_field = True
            if not has_pk_field and not isinstance(model._meta.pk, AutoField):
              # The primary key is not an auto-generated id and it is not mapped in the input...
              header_ok = False
              yield force_text(string_concat(
                model._meta.verbose_name, ': ', _('Missing primary key field %(key)s') % {'key': model._meta.pk.name}
                )) + '\n'
              numerrors += 1
            if not header_ok:
              # Can't process this worksheet
              break
            uploadform = modelform_factory(
              model,
              fields=tuple([i.name for i in headers if isinstance(i, Field)]),
              formfield_callback=lambda f: (isinstance(f, RelatedField) and f.formfield(using=request.database, localize=True)) or f.formfield(localize=True)
              )
          else:
            # Process a data row
            # Step 1: Build a dictionary with all data fields
            d = {}
            colnum = 0
            for cell in row:
              # More fields in data row than headers. Move on to the next row.
              if colnum >= len(headers):
                break
              if isinstance(headers[colnum], Field):
                data = cell.value
                if isinstance(headers[colnum], (IntegerField, AutoField)):
                  if isinstance(data, numericTypes):
                    data = int(data)
                d[headers[colnum].name] = data
              colnum += 1
            # Step 2: Fill the form with data, either updating an existing
            # instance or creating a new one.
            if has_pk_field:
              # A primary key is part of the input fields
              try:
                with transaction.atomic(using=request.database):
                  # Try to find an existing record with the same primary key
                  it = model.objects.using(request.database).get(pk=d[model._meta.pk.name])
                  form = uploadform(d, instance=it)
              except model.DoesNotExist:
                form = uploadform(d)
                it = None
            else:
              # No primary key required for this model
              form = uploadform(d)
              it = None
            # Step 3: Validate the data and save to the database
            if form.has_changed():
              try:
                with transaction.atomic(using=request.database):
                  obj = form.save(commit=False)
                  obj.save(using=request.database)
                  LogEntry(
                    user_id=request.user.pk,
                    content_type_id=contenttype_id,
                    object_id=obj.pk,
                    object_repr=force_text(obj),
                    action_flag=it and CHANGE or ADDITION,
                    change_message=_('Changed %s.') % get_text_list(form.changed_data, _('and'))
                  ).save(using=request.database)
                  if it:
                    changed += 1
                  else:
                    added += 1
              except Exception:
                # Validation fails
                for error in form.non_field_errors():
                  yield force_text(string_concat(
                    model._meta.verbose_name, ': ', _('Row %(rownum)s: %(message)s') % {
                      'rownum': rownum, 'message': error
                    })) + '\n'
                  numerrors += 1
                for field in form:
                  for error in field.errors:
                    yield force_text(string_concat(
                      model._meta.verbose_name, ': ', _('Row %(rownum)s field %(field)s: %(data)s: %(message)s') % {
                        'rownum': rownum, 'data': d[field.name],
                        'field': field.name, 'message': error
                      })) + '\n'
                    numerrors += 1
      # Report status of the import
      yield string_concat(
        model._meta.verbose_name, ": ",
        _('%(rows)d data rows, changed %(changed)d and added %(added)d records, %(errors)d errors') %
          {'rows': rownum - 1, 'changed': changed, 'added': added, 'errors': numerrors}
        ) + '\n'
    yield force_text(_("Done")) + '\n'
