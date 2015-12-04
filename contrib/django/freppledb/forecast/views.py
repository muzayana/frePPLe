#
# Copyright (C) 2012-2013 by frePPLe bvba
#
# All information contained herein is, and remains the property of frePPLe.
# You are allowed to use and modify the source code, as long as the software is used
# within your company.
# You are not allowed to distribute the software, either in the form of source code
# or in the form of compiled binaries.
#

from datetime import datetime
from decimal import Decimal
import json
from openpyxl import load_workbook

from django.conf import settings
from django.contrib.admin.util import unquote
from django.db import connections, transaction
from django.http import HttpResponse, HttpResponseForbidden, Http404
from django.utils import translation, six
from django.utils.formats import get_format
from django.utils.translation import ugettext_lazy as _
from django.utils.text import capfirst
from django.utils.encoding import force_text

from freppledb.boot import getAttributeFields
from freppledb.forecast.models import Forecast, ForecastDemand
from freppledb.common.db import python_date
from freppledb.common.models import BucketDetail
from freppledb.common.report import GridPivot, GridFieldText, GridFieldInteger, GridFieldDate
from freppledb.common.report import EncodedCSVReader, GridReport, GridFieldBool, GridFieldLastModified
from freppledb.common.report import GridFieldChoice, GridFieldNumber, GridFieldDateTime, GridFieldDuration
from freppledb.input.views import PathReport
from freppledb.input.models import Demand, Item, Location, Customer
from freppledb.output.models import Constraint
from freppledb.output.views.constraint import BaseReport


class ForecastList(GridReport):
  '''
  A list report to show forecasts.
  '''
  template = 'admin/base_site_grid.html'
  title = _("forecast")
  basequeryset = Forecast.objects.all()
  model = Forecast
  frozenColumns = 1

  rows = (
    GridFieldText('name', title=_('name'), key=True, formatter='detail', extra="role:'forecast/forecast'"),
    GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail', extra="role:'input/item'"),
    GridFieldText('location', title=_('location'), field_name='location__name', formatter='detail', extra="role:'input/location'"),
    GridFieldText('customer', title=_('customer'), field_name='customer__name', formatter='detail', extra="role:'input/customer'"),
    GridFieldText('description', title=_('description')),
    GridFieldText('category', title=_('category')),
    GridFieldText('subcategory', title=_('subcategory')),
    GridFieldChoice('method', title=_('method'), choices=Forecast.methods),
    GridFieldText('operation', title=_('operation'), field_name='operation__name', formatter='detail', extra="role:'input/operation'"),
    GridFieldInteger('priority', title=_('priority')),
    GridFieldNumber('maxlateness', title=_('maximum lateness')),
    GridFieldNumber('minshipment', title=_('minimum shipment')),
    GridFieldBool('discrete', title=_('discrete')),
    GridFieldBool('planned', title=_('planned')),
    GridFieldText('source', title=_('source')),
    GridFieldLastModified('lastmodified'),
    )


class ForecastDemandList(GridReport):
  '''
  A list report to show forecastdemands.
  '''
  template = 'admin/base_site_grid.html'
  title = _("forecasted demand")
  basequeryset = ForecastDemand.objects.all()
  model = ForecastDemand
  frozenColumns = 1

  rows = (
    GridFieldInteger('id', title=_('identifier'), key=True),
    GridFieldText('forecast', title=_('forecast'), formatter='detail', extra="role:'forecast/forecast'"),
    GridFieldDate('startdate', title=_('start date')),
    GridFieldDate('enddate', title=_('end date')),
    GridFieldNumber('quantity', title=_('quantity')),
    GridFieldText('source', title=_('source')),
    GridFieldLastModified('lastmodified'),
    )


class OverviewReport(GridPivot):
  '''
  A report allowing easy editing of forecast numbers.
  '''
  template = 'forecast/forecast.html'
  title = _('Forecast Report')
  basequeryset = Forecast.objects.all()
  model = Forecast
  permissions = (('view_forecast_report', 'Can view forecast report'),)
  editable = True
  # Limit visibility to months and higher time granularity
  # TODO this should be configurable with a parameter. The parameter should control
  # the time buckets to be used for all forecasting.
  maxBucketLevel = 3
  rows = (
    GridFieldText('forecast', title=_('forecast'), key=True, editable=False, field_name='name', formatter='detail', extra="role:'forecast/forecast'"),
    GridFieldText('item', title=_('item'), field_name='item__name', editable=False, formatter='detail', extra="role:'input/item'"),
    GridFieldText('customer', title=_('customer'), editable=False, field_name='customer__name', formatter='detail', extra="role:'input/customer'"),
    GridFieldText('location', title=_('location'), editable=False, field_name='location__name', formatter='detail', extra="role:'input/location'"),
    GridFieldText('out_method', title=_('selected forecast method'), field_name='out_method', editable=False, hidden=True),
    GridFieldNumber('out_smape', title=_('estimated forecast error'), field_name='out_smape', editable=False, hidden=True)
    )
  crosses = (
    ('orderstotal', {'title': _('total orders')}),
    ('ordersopen', {'title': _('open orders')}),
    ('ordersadjustment', {'title': _('orders adjustment'), 'editable': lambda req: req.user.has_perm('input.change_forecastdemand')}),
    ('forecastbaseline', {'title': _('forecast baseline')}),
    ('forecastadjustment', {'title': _('forecast override'), 'editable': lambda req: req.user.has_perm('input.change_forecastdemand')}),
    ('forecasttotal', {'title': _('forecast total')}),
    ('forecastnet', {'title': _('forecast net')}),
    ('forecastconsumed', {'title': _('forecast consumed')}),
    ('ordersplanned', {'title': _('planned orders')}),
    ('forecastplanned', {'title': _('planned net forecast')}),
    ('past', {'visible': False}),
    ('future', {'visible': False}),
    )

  @classmethod
  def initialize(reportclass, request):
    if reportclass._attributes_added != 2:
      reportclass._attributes_added = 2
      reportclass.attr_sql = ''
      # Adding custom forecast attributes
      for f in getAttributeFields(Forecast, initially_hidden=True):
        reportclass.rows += (f,)
        reportclass.attr_sql += 'forecast.%s, ' % f.name.split('__')[-1]
      # Adding custom item attributes
      for f in getAttributeFields(Item, related_name_prefix="item", initially_hidden=True):
        reportclass.rows += (f,)
        reportclass.attr_sql += 'item.%s, ' % f.name.split('__')[-1]
      # Adding custom location attributes
      for f in getAttributeFields(Location, related_name_prefix="location", initially_hidden=True):
        reportclass.rows += (f,)
        reportclass.attr_sql += 'location.%s, ' % f.name.split('__')[-1]
      # Adding custom customer attributes
      for f in getAttributeFields(Customer, related_name_prefix="customer", initially_hidden=True):
        reportclass.rows += (f,)
        reportclass.attr_sql += 'customer.%s, ' % f.name.split('__')[-1]

  @classmethod
  def extra_context(reportclass, request, *args, **kwargs):
    if args and args[0]:
      return {
        'title': capfirst(force_text(Forecast._meta.verbose_name) + " " + args[0]),
        'post_title': ': ' + capfirst(force_text(_('plan'))),
        }
    else:
      return {}

  @classmethod
  def query(reportclass, request, basequery, sortsql='1 asc'):
    basesql, baseparams = basequery.query.get_compiler(basequery.db).as_sql(with_col_aliases=False)
    # Execute the query
    cursor = connections[request.database].cursor()

    try:
      cursor.execute("SELECT value FROM common_parameter where name='currentdate'")
      d = cursor.fetchone()
      currentdate = datetime.strptime(d[0], "%Y-%m-%d %H:%M:%S").date()
    except:
      currentdate = datetime.now().date()

    # Value or units
    if request.prefs and request.prefs.get('units', 'unit') == 'value':
      suffix = 'value'
    else:
      suffix = ''

    query = '''
        select
          fcstplan.forecast_id,
          forecast.item_id as item_id, forecast.customer_id as customer_id,
          forecast.location_id as location_id, forecast.out_method as out_method,
          forecast.out_smape as out_smape, %s
          fcstplan.bucket as cross1, fcstplan.startdate as cross2, fcstplan.enddate as cross3,
          fcstplan.orderstotal, fcstplan.ordersopen, fcstplan.ordersadjustment,
          fcstplan.forecastbaseline, fcstplan.forecastadjustment, fcstplan.forecasttotal,
          fcstplan.forecastnet, fcstplan.forecastconsumed, fcstplan.ordersplanned,
          fcstplan.forecastplanned
        from (
          select
            fcst.name as forecast_id,
            d.bucket as bucket, d.startdate as startdate, d.enddate as enddate,
            coalesce(sum(forecastplan.orderstotal%s),0) as orderstotal,
            coalesce(sum(forecastplan.ordersopen%s),0) as ordersopen,
            sum(forecastplan.ordersadjustment%s) as ordersadjustment,
            coalesce(sum(forecastplan.forecastbaseline%s),0) as forecastbaseline,
            sum(forecastplan.forecastadjustment%s) as forecastadjustment,
            coalesce(sum(forecastplan.forecasttotal%s),0) as forecasttotal,
            coalesce(sum(forecastplan.forecastnet%s),0) as forecastnet,
            coalesce(sum(forecastplan.forecastconsumed%s),0) as forecastconsumed,
            coalesce(sum(forecastplan.ordersplanned%s),0) as ordersplanned,
            coalesce(sum(forecastplan.forecastplanned%s),0) as forecastplanned
          from (%s) fcst
          -- Multiply with buckets
          cross join (
             select name as bucket, startdate, enddate
             from common_bucketdetail
             where bucket_id = %%s and enddate > %%s and startdate < %%s
             ) d
          -- Forecast plan
          left outer join forecastplan
          on fcst.name = forecastplan.forecast_id
          and forecastplan.startdate >= d.startdate
          and forecastplan.startdate < d.enddate
          -- Grouping
          group by fcst.name,
            d.bucket, d.startdate, d.enddate
          ) fcstplan
        left outer join forecast on
          fcstplan.forecast_id = forecast.name
        left outer join item on
          forecast.item_id = item.name
        left outer join location on
          forecast.location_id = location.name
        left outer join customer on
          forecast.customer_id = customer.name
        order by %s, fcstplan.startdate
        ''' % (
          reportclass.attr_sql,
          suffix, suffix, suffix, suffix, suffix, suffix, suffix, suffix, suffix, suffix,
          basesql, sortsql
          )
    cursor.execute(query, baseparams + (request.report_bucket, request.report_startdate, request.report_enddate) )

    # Build the python result
    for row in cursor.fetchall():
      numfields = len(row)
      res =  {
        'forecast': row[0],
        'item': row[1],
        'customer': row[2],
        'location': row[3],
        'out_method': row[4],
        'out_smape': row[5],
        'bucket': row[numfields-13],
        'startdate': python_date(row[numfields-12]),
        'enddate': python_date(row[numfields-11]),
        'past': python_date(row[numfields-12]) < currentdate and 1 or 0,
        'future': python_date(row[numfields-11]) > currentdate and 1 or 0,
        'orderstotal': row[numfields-10],
        'ordersopen': row[numfields-9],
        'ordersadjustment': row[numfields-8],
        'forecastbaseline': row[numfields-7],
        'forecastadjustment': row[numfields-6],
        'forecasttotal': row[numfields-5],
        'forecastnet': row[numfields-4],
        'forecastconsumed': row[numfields-3],
        'ordersplanned': row[numfields-2],
        'forecastplanned': row[numfields-1],
        }
      # Add attribute fields
      idx = 6
      for f in getAttributeFields(Forecast):
        res[f.field_name] = row[idx]
        idx += 1
      for f in getAttributeFields(Item, related_name_prefix="item"):
        res[f.field_name] = row[idx]
        idx += 1
      for f in getAttributeFields(Location, related_name_prefix="location"):
        res[f.field_name] = row[idx]
        idx += 1
      for f in getAttributeFields(Customer, related_name_prefix="customer"):
        res[f.field_name] = row[idx]
        idx += 1
      yield res


  @classmethod
  def parseJSONupload(reportclass, request):
    # Check permissions
    if not request.user.has_perm('forecast.change_forecastdemand'):
      return HttpResponseForbidden(_('Permission denied'))

    # Loop over the data records
    resp = HttpResponse()
    ok = True
    with transaction.atomic(request.database, savepoint=False):
      for rec in json.JSONDecoder().decode(request.read().decode(request.encoding or settings.DEFAULT_CHARSET)):
        try:
          with transaction.atomic(request.database, savepoint=False):
            fcst = Forecast.objects.all().using(request.database).get(name=rec['id'])
            if rec.get('adjForecast','') != '' or rec.get('adjHistory','') != '':
              fcst.updatePlan(
                datetime.strptime(rec['startdate'], '%Y-%m-%d').date(),
                datetime.strptime(rec['enddate'], '%Y-%m-%d').date(),
                Decimal(rec['adjForecast']) if (rec.get('adjForecast','') != '') else None,
                Decimal(rec['adjHistory']) if (rec.get('adjHistory','') != '') else None,
                True,
                request.database
                )
            else:
              fcst.updatePlan(
                datetime.strptime(rec['startdate'], '%Y-%m-%d').date(),
                datetime.strptime(rec['enddate'], '%Y-%m-%d').date(),
                Decimal(rec['adjForecastValue']) if (rec.get('adjForecastValue','') != '') else None,
                Decimal(rec['adjHistoryValue']) if (rec.get('adjHistoryValue','') != '') else None,
                False,
                request.database
                )

        except Exception as e:
          ok = False
          resp.write(e)
          resp.write('<br/>')
    if ok:
      resp.write("OK")
    resp.status_code = ok and 200 or 403
    return resp


  @classmethod
  def parseCSVupload(reportclass, request):    # TODO also support uploads in pivot format
    # Check permissions
    if not request.user.has_perm('input.change_forecastdemand'):
      yield force_text(_('Permission denied')) + '\n '
    else:

      # Choose the right delimiter and language
      delimiter = ';' if get_format('DECIMAL_SEPARATOR', request.LANGUAGE_CODE, True) == ',' else ','
      if translation.get_language() != request.LANGUAGE_CODE:
        translation.activate(request.LANGUAGE_CODE)

      # Init
      colindexes = [-1, -1, -1, -1, -1, -1, -1]
      rownumber = 0
      prefs = request.user.getPreference(OverviewReport.getKey())
      units = prefs.get('units', 'unit') != 'value' if prefs else True
      pivotbuckets = []

      # Read the name of all buckets in memory
      bucket_names = {
        i.name.lower() : (i.startdate, i.enddate)
        for i in BucketDetail.objects.all().using(request.database).only('name', 'startdate', 'enddate')
        }

      # Handle the complete upload as a single database transaction
      with transaction.atomic(using=request.database):

        # Loop through the data records
        for row in EncodedCSVReader(request.FILES['csv_file'], delimiter=delimiter):
          rownumber += 1

          ### Case 1: The first line is read as a header line
          if rownumber == 1:
            colnum = 0
            for col in row:
              col = col.strip().strip('#').lower()
              if col == _('forecast').lower():
                colindexes[0] = colnum
              elif col == _('start date').lower():
                colindexes[1] = colnum
              elif col == _('end date').lower():
                colindexes[2] = colnum
              elif col == _('bucket').lower():
                colindexes[3] = colnum
              elif col == _('forecast adjustment').lower():
                colindexes[4] = colnum
              elif col == _('orders adjustment').lower():
                colindexes[5] = colnum
              elif col == _("data field").lower():
                # This marks the difference between the pivot and list layout
                colindexes[6] = colnum
              elif colindexes[6] > 0:
                # All columns after 'data field' are assumed to be bucket names
                pivotbuckets.append(col)
              colnum += 1
            errors = False
            if colindexes[0] < 0:
              yield force_text(_('Missing primary key field %(key)s') % {'key': 'forecast'}) + '\n '
              errors = True
            if colindexes[6] > 0:
              if not pivotbuckets:
                yield force_text(_('No time field specified')) + '\n '
                errors = True
            else:
              if colindexes[1] < 0 and colindexes[2] < 0 and colindexes[3] < 0:
                yield force_text(_('No time field specified')) + '\n '
                errors = True
              if colindexes[4] < 0 and colindexes[5] < 0:
                yield force_text(_('No adjustment field specified')) + '\n '
                errors = True
            if errors:
              yield force_text(_('Recognized fields for list layout: forecast, bucket, start date, end date, forecast adjustment, orders adjustment')) + '\n '
              yield force_text(_('Recognized fields for pivot layout: forecast, data field, [bucket names*]')) + '\n '
              break

          ### Case 2: Skip empty rows and comments rows
          elif len(row) == 0 or row[0].startswith('#'):
            continue

          ### Case 3: Process a data row
          else:
            try:
              with transaction.atomic(using=request.database):
                fcstname = row[colindexes[0]] if colindexes[0] < len(row) else None
                fcst = Forecast.objects.all().using(request.database).get(name=fcstname)
                if colindexes[6] > 0:
                  # Case 3a: data row in pivot layout
                  field = row[colindexes[6]].lower() if colindexes[6] > 0 and colindexes[6] < len(row) else None
                  cnt = 0
                  for b in row[colindexes[6] + 1:]:
                    # Loop over data buckets
                    if b:
                      try:
                        (startdate, enddate) = bucket_names[pivotbuckets[cnt]]
                      except KeyError:
                        raise Exception(force_text(_("Bucket '%(name)s' not found") % {'name': pivotbuckets[cnt]}))
                      if field == _('orders adjustment').lower():
                        try:
                          ordersadj = Decimal(b)
                        except:
                          raise Exception(force_text(_("Invalid number: %(value)s") % {'value': b}))
                        fcst.updatePlan(
                          startdate, enddate, None, ordersadj, units, request.database
                          )
                      elif field == _('forecast adjustment').lower():
                        try:
                          fcstadj = Decimal(b)
                        except:
                          raise Exception(force_text(_("Invalid number: %(value)s") % {'value': b}))
                        fcst.updatePlan(
                          startdate, enddate, fcstadj, None, units, request.database
                          )
                    cnt += 1
                else:
                  # Case 3b: data row in list layout
                  startdate = row[colindexes[1]] if colindexes[1] > 0 and colindexes[1] < len(row) else None
                  enddate = row[colindexes[2]] if colindexes[2] > 0 and colindexes[1] < len(row) else None
                  bucket = row[colindexes[3]] if colindexes[3] > 0 and colindexes[1] < len(row) else None
                  if bucket:
                    try:
                      (startdate, enddate) = bucket_names[bucket.lower()]
                    except KeyError:
                      raise Exception(force_text(_("Bucket '%(name)s' not found") % {'name': bucket}))
                  if row[colindexes[5]] and colindexes[4] > 0 and colindexes[1] < len(row):
                    try:
                      fcstadj = Decimal(row[colindexes[4]])
                    except:
                      raise Exception(force_text(_("Invalid number: %(value)s") % {'value': row[colindexes[4]]}))
                  else:
                    fcstadj = None
                  if row[colindexes[5]] and colindexes[5] > 0 and colindexes[5] < len(row):
                    try:
                      ordersadj = Decimal(row[colindexes[5]])
                    except:
                      raise Exception(force_text(_("Invalid number: %(value)s") % {'value': row[colindexes[5]]}))
                  else:
                    ordersadj = None
                  if fcstadj or ordersadj:
                    if not startdate and not enddate:
                      raise Exception(force_text(_("No time bucket specified")))
                    fcst.updatePlan(
                      startdate, enddate, fcstadj, ordersadj, units, request.database
                      )
            except Exception as e:
              yield force_text(
                _('Row %(rownum)s: %(message)s') % {
                  'rownum': rownumber, 'message': str(e)
                  }) + '\n '

      # Report all failed records
      yield force_text(
          _('Uploaded data successfully: processed %d records') % rownumber
          ) + '\n '


  @classmethod
  def parseSpreadsheetUpload(reportclass, request):    # TODO also support uploads in pivot format
    # Check permissions
    if not request.user.has_perm('input.change_forecastdemand'):
      yield force_text(_('Permission denied')) + '\n '
    else:
      # Choose the right language
      if translation.get_language() != request.LANGUAGE_CODE:
        translation.activate(request.LANGUAGE_CODE)

      # Init
      colindexes = [-1, -1, -1, -1, -1, -1, -1]
      rownumber = 0
      prefs = request.user.getPreference(OverviewReport.getKey())
      units = prefs.get('units', 'unit') != 'value' if prefs else True
      pivotbuckets = []

      # Read the name of all buckets in memory
      bucket_names = {
        i.name.lower() : (i.startdate, i.enddate)
        for i in BucketDetail.objects.all().using(request.database).only('name', 'startdate', 'enddate')
        }

      # Handle the complete upload as a single database transaction
      with transaction.atomic(using=request.database):

        # Loop through the data records
        wb = load_workbook(filename=request.FILES['csv_file'], use_iterators=True, data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows():
          rownumber += 1

          ### Case 1: The first line is read as a header line
          if rownumber == 1:
            colnum = 0
            for col in row:
              col = str(col.value).strip().strip('#').lower()
              if col == _('forecast').lower():
                colindexes[0] = colnum
              elif col == _('start date').lower():
                colindexes[1] = colnum
              elif col == _('end date').lower():
                colindexes[2] = colnum
              elif col == _('bucket').lower():
                colindexes[3] = colnum
              elif col == _('forecast adjustment').lower():
                colindexes[4] = colnum
              elif col == _('orders adjustment').lower():
                colindexes[5] = colnum
              elif col == _("data field").lower():
                # This marks the difference between the pivot and list layout
                colindexes[6] = colnum
              elif colindexes[6] > 0:
                # All columns after 'data field' are assumed to be bucket names
                pivotbuckets.append(col)
              colnum += 1
            errors = False
            if colindexes[0] < 0:
              yield force_text(_('Missing primary key field %(key)s') % {'key': 'forecast'}) + '\n '
              errors = True
            if colindexes[6] > 0:
              if not pivotbuckets:
                yield force_text(_('No time field specified')) + '\n '
                errors = True
            else:
              if colindexes[1] < 0 and colindexes[2] < 0 and colindexes[3] < 0:
                yield force_text(_('No time field specified')) + '\n '
                errors = True
              if colindexes[4] < 0 and colindexes[5] < 0:
                yield force_text(_('No adjustment field specified')) + '\n '
                errors = True
            if errors:
              yield force_text(_('Recognized fields for list layout: forecast, bucket, start date, end date, forecast adjustment, orders adjustment')) + '\n '
              yield force_text(_('Recognized fields for pivot layout: forecast, data field, [bucket names*]')) + '\n '
              break

          ### Case 2: Skip empty rows and comments rows
          elif len(row) == 0 or (isinstance(row[0].value, six.string_types) and row[0].value.startswith('#')):
            continue

          ### Case 3: Process a data row
          else:
            try:
              with transaction.atomic(using=request.database):
                fcstname = row[colindexes[0]].value if colindexes[0] < len(row) else None
                fcst = Forecast.objects.all().using(request.database).get(name=fcstname)
                if colindexes[6] > 0:
                  # Case 3a: data row in pivot layout
                  field = row[colindexes[6]].value.lower() if colindexes[6] > 0 and colindexes[6] < len(row) else None
                  cnt = 0
                  for b in row[colindexes[6] + 1:]:
                    # Loop over data buckets
                    if b.value is not None:
                      try:
                        (startdate, enddate) = bucket_names[pivotbuckets[cnt]]
                      except KeyError:
                        raise Exception(force_text(_("Bucket '%(name)s' not found") % {'name': pivotbuckets[cnt]}))
                      if field == _('orders adjustment').lower():
                        try:
                          ordersadj = Decimal(b.value)
                        except:
                          raise Exception(force_text(_("Invalid number: %(value)s") % {'value': b.value}))
                        fcst.updatePlan(
                          startdate, enddate, None, ordersadj, units, request.database
                          )
                      elif field == _('forecast adjustment').lower():
                        try:
                          fcstadj = Decimal(b.value)
                        except:
                          raise Exception(force_text(_("Invalid number: %(value)s") % {'value': b.value}))
                        fcst.updatePlan(
                          startdate, enddate, fcstadj, None, units, request.database
                          )
                    cnt += 1
                else:
                  # Case 3b: data row in list layout
                  startdate = row[colindexes[1]].value if colindexes[1] > 0 and colindexes[1] < len(row) else None
                  enddate = row[colindexes[2]].value if colindexes[2] > 0 and colindexes[2] < len(row) else None
                  bucket = row[colindexes[3]].value if colindexes[3] > 0 and colindexes[3] < len(row) else None
                  if bucket:
                    try:
                      (startdate, enddate) = bucket_names[bucket.lower()]
                    except KeyError:
                      raise Exception(force_text(_("Bucket '%(name)s' not found") % {'name': bucket}))
                  if row[colindexes[4]].value is not None and colindexes[4] > 0 and colindexes[4] < len(row):
                    try:
                      fcstadj = Decimal(row[colindexes[4]].value)
                    except:
                      raise Exception(force_text(_("Invalid number: %(value)s") % {'value': row[colindexes[4]].value}))
                  else:
                    fcstadj = None
                  if row[colindexes[5]].value is not None and colindexes[5] > 0 and colindexes[5] < len(row):
                    try:
                      ordersadj = Decimal(row[colindexes[5]].value)
                    except:
                      raise Exception(force_text(_("Invalid number: %(value)s") % {'value': row[colindexes[5]].value}))
                  else:
                    ordersadj = None
                  if fcstadj or ordersadj:
                    if not startdate and not enddate:
                      raise Exception(force_text(_("No time bucket specified")))
                    fcst.updatePlan(
                      startdate, enddate, fcstadj, ordersadj, units, request.database
                      )
            except Exception as e:
              yield force_text(
                _('Row %(rownum)s: %(message)s') % {
                  'rownum': rownumber, 'message': e
                  }) + '\n '

      # Report all failed records
      yield force_text(
          _('Uploaded data successfully: processed %d records') % rownumber
          ) + '\n '


class UpstreamForecastPath(PathReport):
  downstream = False
  objecttype = Forecast

  @classmethod
  def getRoot(reportclass, request, entity):
    from django.core.exceptions import ObjectDoesNotExist

    try:
      fcst = Forecast.objects.using(request.database).get(name=entity)
    except ObjectDoesNotExist:
      raise Http404("forecast %s doesn't exist" % entity)

    if fcst.operation:
      # Delivery operation on the forecast
      return [ (0, None, fcst.operation, 1, 0, None, 0, False) ]
    elif fcst.item.operation:
      # Delivery operation on the item
      return [ (0, None, fcst.item.operation, 1, 0, None, 0, False) ]
    else:
      # Autogenerated delivery operation
      result = reportclass.findDeliveries(fcst.item, fcst.location, request.database)
      if result:
        return result
      else:
        raise Http404("No supply path defined for forecast %s" % entity)


class OrderReport(GridReport):
  '''
  A list report to show demands.
  '''
  template = 'input/demandlist.html'
  title = _("Demand List")
  model = Demand
  frozenColumns = 1

  @ classmethod
  def basequeryset(reportclass, request, args, kwargs):
    fcst = Forecast.objects.using(request.database).get(name__exact=unquote(request.GET['forecast']))
    return Demand.objects.filter(
      item__lft__gte=fcst.item.lft,
      item__lft__lt=fcst.item.rght,
      customer__lft__gte=fcst.customer.lft,
      customer__lft__lt=fcst.customer.rght
      )

  rows = (
    GridFieldText('name', title=_('name'), key=True, formatter='detail', extra="role:'input/demand'"),
    GridFieldText('item', title=_('item'), field_name='item__name', formatter='detail', extra="role:'input/item'"),
    GridFieldText('customer', title=_('customer'), field_name='customer__name', formatter='detail', extra="role:'input/customer'"),
    GridFieldText('description', title=_('description')),
    GridFieldText('category', title=_('category')),
    GridFieldText('subcategory', title=_('subcategory')),
    GridFieldDateTime('due', title=_('due')),
    GridFieldNumber('quantity', title=_('quantity')),
    GridFieldText('operation', title=_('delivery operation'), formatter='detail', extra="role:'input/operation'"),
    GridFieldInteger('priority', title=_('priority')),
    GridFieldText('owner', title=_('owner'), formatter='detail', extra="role:'input/demand'"),
    GridFieldChoice('status', title=_('status'), choices=Demand.demandstatus),
    GridFieldDuration('maxlateness', title=_('maximum lateness')),
    GridFieldNumber('minshipment', title=_('minimum shipment')),
    GridFieldText('source', title=_('source')),
    GridFieldLastModified('lastmodified'),
    )


class ConstraintReport(BaseReport):

  template = 'forecast/constraint_forecast.html'

  @ classmethod
  def basequeryset(reportclass, request, args, kwargs):
    if args and args[0]:
      request.session['lasttab'] = 'constraint'
      return Constraint.objects.all().filter(demand__startswith=args[0])
    else:
      return Constraint.objects.all()
